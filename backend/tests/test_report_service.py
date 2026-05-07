"""
test_report_service.py — Tests for Excel report export.

Run: pytest backend/tests/test_report_service.py -v
"""

import asyncio
import io
from datetime import datetime, timezone
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import load_workbook

from app.services import report_service


def _scalars_result(items):
    return SimpleNamespace(scalars=lambda: SimpleNamespace(all=lambda: items))


def _rows_result(rows):
    return SimpleNamespace(all=lambda: rows)


class FakeDb:
    def __init__(self, results):
        self._results = list(results)

    async def execute(self, _stmt):
        return self._results.pop(0)


def test_generate_report_creates_expected_three_sheet_workbook(monkeypatch):
    session_id = uuid4()
    timestamp = datetime(2026, 4, 7, 9, 1, tzinfo=timezone.utc)
    cafe = SimpleNamespace(cafe_id=1, name="Cafe A")
    score = SimpleNamespace(
        total_sessions=3,
        behavior_score=7.1,
        has_enough_data=False,
    )

    db = FakeDb([
        _rows_result([
            SimpleNamespace(
                session_id=session_id,
                device_id="device-001",
                cafe="Cafe A",
                start_time=timestamp,
                end_time=datetime(2026, 4, 7, 11, 1, tzinfo=timezone.utc),
                duration_min=120.0,
                gps_log_count=2,
                status="completed",
            )
        ]),
        _rows_result([
            SimpleNamespace(
                session_id=session_id,
                timestamp=timestamp,
                lat=21.0285,
                lng=105.8542,
                accuracy=12.5,
                cafe_id=1,
            )
        ]),
        _scalars_result([cafe]),
        _rows_result([(1, 3, 95.25)]),
    ])

    async def fake_get_latest_scores_by_cafe_id(_db, cafe_ids):
        assert cafe_ids == [1]
        return {1: score}

    monkeypatch.setattr(
        report_service,
        "get_latest_scores_by_cafe_id",
        fake_get_latest_scores_by_cafe_id,
    )

    output = asyncio.run(report_service.generate_report(db))
    wb = load_workbook(io.BytesIO(output.getvalue()))

    assert wb.sheetnames == ["Sessions", "GPS Logs", "Cafe Summary"]
    assert [cell.value for cell in wb["Sessions"][1]] == [
        "session_id",
        "device_id",
        "cafe",
        "start_time",
        "end_time",
        "duration_min",
        "gps_log_count",
        "status",
    ]
    assert [cell.value for cell in wb["GPS Logs"][1]] == [
        "session_id",
        "timestamp",
        "lat",
        "lng",
        "accuracy",
        "cafe_id",
    ]
    assert [cell.value for cell in wb["Cafe Summary"][1]] == [
        "cafe",
        "total_sessions",
        "avg_duration",
        "behavior_score",
        "has_enough_data",
    ]

    assert wb["Sessions"].cell(row=2, column=1).value == str(session_id)
    assert wb["Sessions"].cell(row=2, column=3).value == "Cafe A"
    assert wb["Sessions"].cell(row=2, column=7).value == 2
    assert wb["GPS Logs"].cell(row=2, column=5).value == 12.5
    assert wb["Cafe Summary"].cell(row=2, column=1).value == "Cafe A"
    assert wb["Cafe Summary"].cell(row=2, column=2).value == 3
    assert wb["Cafe Summary"].cell(row=2, column=3).value == 95.2
    assert wb["Cafe Summary"].cell(row=2, column=4).value == 7.1


def test_generate_report_falls_back_to_session_count_when_score_missing(monkeypatch):
    cafe = SimpleNamespace(cafe_id=1, name="Cafe A")

    db = FakeDb([
        _rows_result([]),
        _rows_result([]),
        _scalars_result([cafe]),
        _rows_result([(1, 2, 72.75)]),
    ])

    async def fake_get_latest_scores_by_cafe_id(_db, cafe_ids):
        assert cafe_ids == [1]
        return {}

    monkeypatch.setattr(
        report_service,
        "get_latest_scores_by_cafe_id",
        fake_get_latest_scores_by_cafe_id,
    )

    output = asyncio.run(report_service.generate_report(db))
    wb = load_workbook(io.BytesIO(output.getvalue()))
    ws = wb["Cafe Summary"]

    assert ws.cell(row=2, column=2).value == 2
    assert ws.cell(row=2, column=3).value == 72.8
    assert ws.cell(row=2, column=5).value is False
