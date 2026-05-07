"""
report_service.py — Business Logic: Report/Export.

Tổng hợp dữ liệu session/GPS/cafe score, tạo file Excel (.xlsx).
"""

import io
from datetime import datetime

from openpyxl import Workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.cafe import Cafe
from app.models.gps_log import GpsLog
from app.models.session import Session as StudySession
from app.services.cafe_score_service import get_latest_scores_by_cafe_id


def _format_excel_timestamp(value: datetime | None) -> str | None:
    if value is None:
        return None
    return value.isoformat()


async def _get_session_rows(db: AsyncSession):
    stmt = (
        select(
            StudySession.session_id,
            StudySession.device_id,
            Cafe.name.label("cafe"),
            StudySession.start_time,
            StudySession.end_time,
            StudySession.duration_min,
            func.count(GpsLog.log_id).label("gps_log_count"),
            StudySession.status,
        )
        .outerjoin(Cafe, StudySession.cafe_id == Cafe.cafe_id)
        .outerjoin(GpsLog, StudySession.session_id == GpsLog.session_id)
        .group_by(
            StudySession.session_id,
            StudySession.device_id,
            Cafe.name,
            StudySession.start_time,
            StudySession.end_time,
            StudySession.duration_min,
            StudySession.status,
        )
        .order_by(StudySession.start_time.desc())
    )
    result = await db.execute(stmt)
    return result.all()


async def _get_gps_log_rows(db: AsyncSession):
    stmt = (
        select(
            GpsLog.session_id,
            GpsLog.timestamp,
            GpsLog.lat,
            GpsLog.lng,
            GpsLog.accuracy_m.label("accuracy"),
            StudySession.cafe_id,
        )
        .join(StudySession, GpsLog.session_id == StudySession.session_id)
        .order_by(GpsLog.timestamp.asc(), GpsLog.log_id.asc())
    )
    result = await db.execute(stmt)
    return result.all()


async def _get_active_cafes(db: AsyncSession):
    stmt = select(Cafe).where(Cafe.status == "active").order_by(Cafe.cafe_id.asc())
    result = await db.execute(stmt)
    return result.scalars().all()


async def _get_session_summary_by_cafe_id(db: AsyncSession) -> dict[int, dict[str, object]]:
    stmt = (
        select(
            StudySession.cafe_id,
            func.count(StudySession.session_id).label("total_sessions"),
            func.avg(StudySession.duration_min).label("avg_duration"),
        )
        .where(
            StudySession.cafe_id.is_not(None),
            StudySession.status == "completed",
        )
        .group_by(StudySession.cafe_id)
    )
    result = await db.execute(stmt)

    return {
        cafe_id: {
            "total_sessions": int(total_sessions or 0),
            "avg_duration": round(float(avg_duration), 1)
            if avg_duration is not None
            else None,
        }
        for cafe_id, total_sessions, avg_duration in result.all()
        if cafe_id is not None
    }


async def generate_report(db: AsyncSession) -> io.BytesIO:
    """Tổng hợp dữ liệu và tạo file Excel gồm Sessions, GPS Logs, Cafe Summary."""
    session_rows = await _get_session_rows(db)
    gps_log_rows = await _get_gps_log_rows(db)
    cafes = await _get_active_cafes(db)
    cafe_ids = [cafe.cafe_id for cafe in cafes]
    scores_by_cafe_id = await get_latest_scores_by_cafe_id(db, cafe_ids)
    session_summary_by_cafe_id = await _get_session_summary_by_cafe_id(db)

    wb = Workbook()

    sessions_ws = wb.active
    sessions_ws.title = "Sessions"
    sessions_ws.append([
        "session_id",
        "device_id",
        "cafe",
        "start_time",
        "end_time",
        "duration_min",
        "gps_log_count",
        "status",
    ])
    for row in session_rows:
        sessions_ws.append([
            str(row.session_id),
            row.device_id,
            row.cafe,
            _format_excel_timestamp(row.start_time),
            _format_excel_timestamp(row.end_time),
            row.duration_min,
            row.gps_log_count,
            row.status,
        ])

    gps_ws = wb.create_sheet("GPS Logs")
    gps_ws.append([
        "session_id",
        "timestamp",
        "lat",
        "lng",
        "accuracy",
        "cafe_id",
    ])
    for row in gps_log_rows:
        gps_ws.append([
            str(row.session_id),
            _format_excel_timestamp(row.timestamp),
            row.lat,
            row.lng,
            row.accuracy,
            row.cafe_id,
        ])

    cafe_ws = wb.create_sheet("Cafe Summary")
    cafe_ws.append([
        "cafe",
        "total_sessions",
        "avg_duration",
        "behavior_score",
        "has_enough_data",
    ])
    for cafe in cafes:
        score = scores_by_cafe_id.get(cafe.cafe_id)
        session_summary = session_summary_by_cafe_id.get(cafe.cafe_id, {})
        total_sessions = (
            score.total_sessions
            if score and score.total_sessions is not None
            else session_summary.get("total_sessions")
        )
        cafe_ws.append([
            cafe.name,
            total_sessions,
            session_summary.get("avg_duration"),
            score.behavior_score if score else None,
            score.has_enough_data if score else False,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
